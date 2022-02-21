"""
excel类需要有以下功能
1.打开excel文件
2.读取sheet表单
3.读取所有数据
4.指定单元格写入数据（使用静态方法，不要使用实例方法）
"""

import openpyxl
from openpyxl.worksheet import worksheet
from openpyxl.worksheet.worksheet import Worksheet


class ExcelHandler():
    """操作excel"""

    def __init__(self,file):
        """初始化函数,文件路径名称作为属性，不将sheet_name作为属性，不然每次只能打开一个sheet"""
        self.file = file

    def open_sheet(self,sheet_name) -> Worksheet:
        """打开表单,返回的是一个worksheet对象,加->worksheet注解后，sheet.可以有提示"""
        wb = openpyxl.load_workbook(self.file)
        sheet = wb[sheet_name]
        wb.close()
        return sheet

    def header(self,sheet_name):
        """获取表单的表头，第一行数据"""
        # 得到的是一个sheet对象
        sheet = self.open_sheet(sheet_name)
        # 使用列表存表头的数据
        headers = []
        for i in sheet[1]:
              headers.append(i.value)
        return headers

    def read(self,sheet_name):
        """读取所有的数据"""
        sheet = self.open_sheet(sheet_name)
        # sheet.rows得到的是一个生成器，还没学，用list转换
        rows = list(sheet.rows)
        data = []
        # 用一个切片将第一行剔除掉
        for row in rows[1:]:
            # 一行的所有数据存储到列表row_data中
            row_data = []
            for cell in row:
                # print(cell.value)
                row_data.append(cell.value)
            # 转成字典，和header进行zip方法打包，会是zip对象，一一对应后转化为字典
            data_dict = dict(zip(self.header(sheet_name),row_data))
            # 每一行的数据存储到列表data中
            data.append(data_dict)
        return data
    # 用实例方法容易出问题，如果文件不关闭，会造成数据的紊乱
    # 如果将wb=load_workbook(file)写在init初始化方法中，每次打开的都是初始化的表
    # 所以第一次修改后再修改还是初始化的表
    # 避免出现：每次操作表单时都去使用一个新的workbook
    # 使用实例方法，不把wb放到初始化实例方法中
    def write(self,sheet_name,row,col,new_value):
        """写入Excel数据方法"""
        wb = openpyxl.load_workbook(self.file)
        sheet = wb[sheet_name]
        # 对单元格写入新的值
        sheet.cell(row,col).value = new_value
        # 保存workbook
        wb.save(self.file)
        wb.close()


    # 静态方法实现，跟实例方法传递的方法差不多
    # @staticmethod
    # def write(file,sheet_name,row,col,new_data):
    #     wb = openpyxl.load_workbook(file)
    #     sheet = wb[sheet_name]
    #     sheet.cell(row,col).value = new_data
        #报错wokebook文件
    #     wb.save(file)
    #     wb.close()








if __name__ == '__main__':
    excel = ExcelHandler(r"C:\Users\86176\Desktop\cases.xlsx")
    # data = excel.write('Sheet1',1,1,'username')
    data = excel.read('Sheet1')
    # excel.write(r"C:\Users\86176\Desktop\cases.xlsx",'Sheet1',1,1,'3333')
    print(data)

    # zip ['title','gender'] [1,2]，将列表打包成字典