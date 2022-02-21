"""
保存测试数据
Excel的基础用法，手工如何操作Excel，python学习手工如何操作Excel

手工操作Excel的流程
1.打开excel文件（路径+文件名）
2.获取表单
3.使用行号和列号确定需要读取的数据
4.关闭文件

python操作excel，工具
-openpyxl, 支持xlsx新型格式的读写，读取速度还可以，老格式xls不支持
-tablib（推荐）, 支持多种格式读写，xlsx,xls,csv,json,yaml,html,pd
-xlrd, 经典的excel的读取库，不支持写
-pandas，功能强，太臃肿了

安装 pip install openpyxl

获取表单：
1.wb.active 获取被选中，激活的表单
2.通过索引，wb.worksheets[索引]
3.通过sheet名字，wb["sheet名字"]
"""

import openpyxl
# 1.读取excel的文件，获得一个workbook对象：
# load_workbook(filename, read_only=False, keep_vba=KEEP_VBA,data_only=False, keep_links=True)
# 只读时可以提升读取速度，read_only=True
# Windows下面的路径有反斜杠\，加r，不被转义
# 读取时一定要关闭该文件，不然无法读取,如果是.xls改为xlsx的也无法读取，需新建
wb = openpyxl.load_workbook(r'C:\Users\86176\Desktop\cases.xlsx')
# print(wb)
# # 不直接去获取worksheets属性，称为私有属性
# print(wb._sheets) #self._sheets

# active是表示被激活，被选择的sheet
# active_sheet = wb.active

# sheetnames和_sheets有什么区别？
# sheetnames列表中存储的是字符串，_sheets里面表示的是对象

# 获取所有表单的正确用法
wb.worksheets
# 获取某一个表单，1.通过索引去获取
# sheet = wb.worksheets[0]
# print(sheet)
# 通过表单名字去获取
# 旧方法
# sheet = wb.get_sheet_by_name("Sheet1")
# print(sheet)
# 正规用法
sheet = wb["Sheet1"]
# print(sheet)
# pycharm不支持sheet.属性的提示，索引获取sheet后可以sheet.属性

# 2.读取单个单元格,行和列
# 行和列是从1开始的，不是python当中从0开始,得到的是一个对象
# cell = sheet.cell(1,2)
# print(cell)
# # 获取cell的值,很多同学会忘记加上value
# print(cell.value)
#
# # 获取某一行
# print(sheet[1])
# # 获取一行所有值
# for col in sheet[1]:
#     print(col.value)
# # 获取某一列
# print(sheet['A'])
# # 如何获取多行,1-3行，第三行是包含的
# print(sheet[1:3])

# 获取所有的数据,sheet.rows得到的是一个worksheet对象，转换为一个列表
total_data = list(sheet.rows)
print(sheet.rows)
print(total_data)

for row in total_data:
    for cell in row:
        print(cell.value)

# 写入：一个单元格
# 保存,save("文件名称")
wb.save(r'C:\Users\86176\Desktop\cases.xlsx')
# 关闭
wb.close()



wb = openpyxl.load_workbook(r'C:\Users\86176\Desktop\cases.xlsx')
sheet = wb['Sheet1']
row1 = sheet[1]

col1 = sheet['A']
cell1 =sheet.cell(1,1)
print(cell1.value)
print(row1[0])